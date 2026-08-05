"""
共享工具模块 — 问题1-5 核心计算
================================
泛化为多导弹、多弹版本。

功能:
  - compute_T_eff(Ad, td, missile)        单弹单导弹遮蔽时长
  - compute_T_eff_intervals(Ad, td, missile)  遮蔽区间列表
  - multi_bomb_T_eff(bombs, missile)        多弹并集遮蔽时长
  - union_duration(intervals_list)          区间并集总长
  - drone_feasible(Ad, td, drone_start, dz) 无人机可行性
  - save_result_xlsx(result, filename)      xlsx输出
"""

import numpy as np

# 纯Python二分法求根 — 替代 scipy.optimize.bisect，避免C层segfault
def _bisect_py(f, a, b, xtol=1e-6, maxiter=50):
    """Pure-Python bisection root-finding. Safe against NaN/inf."""
    fa, fb = f(a), f(b)
    if fa is None or fb is None:
        raise ValueError("function returned None")
    if not np.isfinite(fa) or not np.isfinite(fb):
        raise ValueError("function returned non-finite value")
    if fa * fb >= 0:
        raise ValueError("f(a) and f(b) must have opposite signs")
    for _ in range(maxiter):
        c = (a + b) / 2.0
        if abs(b - a) < xtol:
            return c
        fc = f(c)
        if fc is None or not np.isfinite(fc):
            return c  # fallback
        if fa * fc < 0:
            b, fb = c, fc
        else:
            a, fa = c, fc
    return (a + b) / 2.0

# ======================== 场景常量 ========================
M1_0  = np.array([20000., 0., 2000.])
M2_0  = np.array([19000., 600., 2100.])
M3_0  = np.array([18000., -600., 1900.])

FY1_0 = np.array([17800., 0., 1800.])
FY2_0 = np.array([12000., 1400., 1400.])
FY3_0 = np.array([6000., -3000., 700.])
FY4_0 = np.array([11000., 2000., 1800.])
FY5_0 = np.array([13000., -2000., 1300.])

DRONE_POS = {'FY1': FY1_0, 'FY2': FY2_0, 'FY3': FY3_0,
             'FY4': FY4_0, 'FY5': FY5_0}

O_DECOY = np.array([0., 0., 0.])
O1 = np.array([0., 200., 0.])

Rc, Hc = 7., 10.
R_SMOKE = 10.
VS = 3.
TL = 20.
VM = 300.
G = 9.8

V_MIN, V_MAX = 70., 140.

# 导弹飞行时间: 从起点到假目标原点的总飞行时间
MISSILE_T_FLIGHT = {
    'M1': np.linalg.norm(M1_0 - O_DECOY) / VM,  # ≈ 67.0s
    'M2': np.linalg.norm(M2_0 - O_DECOY) / VM,  # ≈ 63.8s
    'M3': np.linalg.norm(M3_0 - O_DECOY) / VM,  # ≈ 60.4s
}

u_m1 = (O_DECOY - M1_0) / np.linalg.norm(O_DECOY - M1_0)
u_m2 = (O_DECOY - M2_0) / np.linalg.norm(O_DECOY - M2_0)
u_m3 = (O_DECOY - M3_0) / np.linalg.norm(O_DECOY - M3_0)

MISSILE_PARAMS = {
    'M1': {'M0': M1_0, 'u_m': u_m1},
    'M2': {'M0': M2_0, 'u_m': u_m2},
    'M3': {'M0': M3_0, 'u_m': u_m3},
}


def M_fn(t, missile='M1'):
    p = MISSILE_PARAMS[missile]
    return p['M0'] + p['u_m'] * VM * t


def sample_boundary(Mp):
    Mx, My, Mz = Mp
    Ox, Oy, Oz = O1
    if not (np.isfinite(Mx) and np.isfinite(My) and np.isfinite(Mz)):
        return np.empty((0, 3))
    d_xy = np.array([Mx - Ox, My - Oy])
    dist_xy = np.linalg.norm(d_xy)
    if dist_xy < 1e-6:
        return np.empty((0, 3))
    u_xy = d_xy / dist_xy
    psi = np.arcsin(np.clip(Rc / dist_xy, 0, 1))
    cp, sp = np.cos(psi), np.sin(psi)
    tR = np.array([u_xy[0] * cp - u_xy[1] * sp, u_xy[0] * sp + u_xy[1] * cp])
    tL = np.array([u_xy[0] * cp + u_xy[1] * sp, -u_xy[0] * sp + u_xy[1] * cp])
    TR = np.array([Ox + Rc * tR[0], Oy + Rc * tR[1]])
    TL = np.array([Ox + Rc * tL[0], Oy + Rc * tL[1]])
    pts = []
    ns, ne = 20, 48
    for i in range(ns + 1):
        z = Oz + Hc * i / ns
        pts.append(np.array([TL[0], TL[1], z]))
        pts.append(np.array([TR[0], TR[1], z]))
    for i in range(ne + 1):
        a = 2 * np.pi * i / ne
        pts.append(np.array([Ox + Rc * np.cos(a), Oy + Rc * np.sin(a), Oz + Hc]))
    aM = np.arctan2(My - Oy, Mx - Ox)
    for i in range(ne // 2 + 1):
        a = aM + np.pi / 2 + np.pi * i / (ne // 2)
        pts.append(np.array([Ox + Rc * np.cos(a), Oy + Rc * np.sin(a), Oz]))
    for i in range(1, 10):
        f = i / 11.
        Tm = TL + f * (TR - TL)
        v = Tm - np.array([Ox, Oy])
        vn = np.linalg.norm(v)
        if vn < 1e-10:
            continue
        v = v * (Rc / vn)
        Tm = np.array([Ox, Oy]) + v
        for j in range(ns // 2 + 1):
            z = Oz + Hc * j / (ns // 2)
            pts.append(np.array([Tm[0], Tm[1], z]))
    vecs = []
    for p in pts:
        v = p - Mp
        d = np.linalg.norm(v)
        if d > 1e-10 and np.isfinite(d):
            vecs.append(v / d)
    if not vecs:
        return np.empty((0, 3))
    return np.array(vecs)


def th_al(t, Ad, td, missile='M1'):
    Mt = M_fn(t, missile)
    age = t - td
    if age < 0 or age > TL:
        return None, None
    At = Ad - np.array([0., 0., VS * age])
    MA = np.linalg.norm(Mt - At)
    if MA <= R_SMOKE:
        return 0.0, np.pi / 2
    u_A = (At - Mt) / MA
    alpha = np.arcsin(R_SMOKE / MA)
    proj = sample_boundary(Mt)
    if len(proj) == 0:
        return None, None
    dots = np.clip(proj @ u_A, -1, 1)
    return np.max(np.arccos(dots)), alpha


def compute_T_eff(Ad, td, missile='M1'):
    p = MISSILE_PARAMS[missile]
    M0 = p['M0']
    u_m = p['u_m']

    def f_C2(t):
        th, al = th_al(t, Ad, td, missile)
        if th is None:
            return -1
        return al - th

    t_scan = np.linspace(td, td + TL, 40)
    c2_roots = []
    for i in range(len(t_scan) - 1):
        f1 = f_C2(t_scan[i])
        f2 = f_C2(t_scan[i + 1])
        if f1 is not None and f2 is not None and f1 * f2 < 0:
            try:
                r = _bisect_py(f_C2, t_scan[i], t_scan[i + 1], xtol=1e-6, maxiter=50)
                c2_roots.append(r)
            except:
                pass

    dx0 = M0[0] - Ad[0]
    dy0 = M0[1] - Ad[1]
    dz0 = M0[2] - Ad[2] - VS * td
    vx = u_m[0] * VM
    vy = u_m[1] * VM
    vz = u_m[2] * VM + VS
    a2 = vx**2 + vy**2 + vz**2
    a1 = 2 * (dx0 * vx + dy0 * vy + dz0 * vz)
    a0 = dx0**2 + dy0**2 + dz0**2 - R_SMOKE**2
    disc = a1**2 - 4 * a2 * a0
    c1_roots = []
    if disc >= 0:
        sqrt_disc = np.sqrt(disc)
        for tc in [(-a1 - sqrt_disc) / (2 * a2), (-a1 + sqrt_disc) / (2 * a2)]:
            if td < tc < td + TL:
                c1_roots.append(tc)
        c1_roots.sort()

    v_mx, v_my, v_mz = u_m[0] * VM, u_m[1] * VM, u_m[2] * VM
    Az_c = Ad[2] + VS * td
    a2g = (v_mx**2 + v_my**2 + v_mz**2) - VS**2
    a1g = 2 * (M0[0] * v_mx + (M0[1] - O1[1]) * v_my + M0[2] * v_mz + Az_c * VS)
    a0g = (M0[0]**2 + (M0[1] - O1[1])**2 + M0[2]**2) - (Ad[0]**2 + (Ad[1] - O1[1])**2 + Az_c**2)
    discg = a1g**2 - 4 * a2g * a0g
    t_geom = td + TL + 999
    if discg >= 0:
        sqrt_discg = np.sqrt(discg)
        for tg in [(-a1g - sqrt_discg) / (2 * a2g), (-a1g + sqrt_discg) / (2 * a2g)]:
            if td < tg < td + TL:
                t_geom = min(t_geom, tg)

    t_end = td + TL
    t_flight_end = MISSILE_T_FLIGHT.get(missile, 1e9)
    if t_end > t_flight_end:
        t_end = t_flight_end

    # Evaluate initial C2/C1 state
    f0 = f_C2(td)
    c2_on = (f0 is not None and f0 >= 0)
    # C1 initial: is missile already inside smoke sphere?
    Mt0 = M_fn(td, missile)
    At0 = Ad - np.array([0., 0., VS * 0])
    MA0 = np.linalg.norm(Mt0 - At0)
    c1_on = (MA0 <= R_SMOKE)

    # Pair C2 roots into [start, end] intervals
    c2_pairs = []
    c2_idx = 0
    cur_start = td
    active = c2_on
    while c2_idx < len(c2_roots):
        if active:
            # Currently on, root switches off
            end_t = c2_roots[c2_idx]
            if end_t > cur_start:
                c2_pairs.append((cur_start, end_t))
            active = False
            cur_start = c2_roots[c2_idx]
        else:
            # Currently off, root switches on
            cur_start = c2_roots[c2_idx]
            active = True
        c2_idx += 1
    if active:
        # Still on after all roots
        end_t = t_end
        if end_t > cur_start:
            c2_pairs.append((cur_start, end_t))

    # Pair C1 roots similarly
    c1_pairs = []
    c1_idx = 0
    cur_start = td
    active = c1_on
    while c1_idx < len(c1_roots):
        if active:
            end_t = c1_roots[c1_idx]
            if end_t > cur_start:
                c1_pairs.append((cur_start, end_t))
            active = False
            cur_start = c1_roots[c1_idx]
        else:
            cur_start = c1_roots[c1_idx]
            active = True
        c1_idx += 1
    if active:
        end_t = t_end
        if end_t > cur_start:
            c1_pairs.append((cur_start, end_t))

    # Merge: C2 needs geometry constraint (MO1 > AO1), C1 doesn't
    intervals = []
    for (s, e) in c2_pairs:
        # C2 only valid when smoke is between missile and target
        valid_end = min(e, t_geom, t_end)
        if valid_end > s:
            intervals.append((s, valid_end, 'C2'))
    for (s, e) in c1_pairs:
        # C1: missile inside smoke sphere, no geometry constraint
        if e > s:
            intervals.append((s, e, 'C1'))

    if not intervals:
        dur = 0.
        for tt in np.arange(td, t_end, 0.1):
            Mt = M_fn(tt, missile)
            age = tt - td
            At = Ad - np.array([0., 0., VS * age])
            MA = np.linalg.norm(Mt - At)
            if MA <= R_SMOKE:
                dur += 0.1
                continue
            MO1 = np.linalg.norm(Mt - O1)
            AO1 = np.linalg.norm(At - O1)
            if MO1 <= AO1:
                continue
            th, al = th_al(tt, Ad, td, missile)
            if th is not None and th <= al:
                dur += 0.1
        return dur
    return sum(e - s for s, e, _ in intervals)


def compute_T_eff_intervals(Ad, td, missile='M1'):
    """返回遮蔽区间列表 [(t_start, t_end), ...]"""
    p = MISSILE_PARAMS[missile]
    M0 = p['M0']
    u_m = p['u_m']

    def f_C2(t):
        th, al = th_al(t, Ad, td, missile)
        if th is None:
            return -1
        return al - th

    t_scan = np.linspace(td, td + TL, 40)
    c2_roots = []
    for i in range(len(t_scan) - 1):
        f1 = f_C2(t_scan[i])
        f2 = f_C2(t_scan[i + 1])
        if f1 is not None and f2 is not None and f1 * f2 < 0:
            try:
                r = _bisect_py(f_C2, t_scan[i], t_scan[i + 1], xtol=1e-6, maxiter=50)
                c2_roots.append(r)
            except:
                pass

    dx0 = M0[0] - Ad[0]; dy0 = M0[1] - Ad[1]; dz0 = M0[2] - Ad[2] - VS * td
    vx = u_m[0] * VM; vy = u_m[1] * VM; vz = u_m[2] * VM + VS
    a2 = vx**2 + vy**2 + vz**2
    a1 = 2 * (dx0 * vx + dy0 * vy + dz0 * vz)
    a0 = dx0**2 + dy0**2 + dz0**2 - R_SMOKE**2
    disc = a1**2 - 4 * a2 * a0
    c1_roots = []
    if disc >= 0:
        sqrt_disc = np.sqrt(disc)
        for tc in [(-a1 - sqrt_disc) / (2 * a2), (-a1 + sqrt_disc) / (2 * a2)]:
            if td < tc < td + TL: c1_roots.append(tc)
        c1_roots.sort()

    v_mx, v_my, v_mz = u_m[0] * VM, u_m[1] * VM, u_m[2] * VM
    Az_c = Ad[2] + VS * td
    a2g = (v_mx**2 + v_my**2 + v_mz**2) - VS**2
    a1g = 2 * (M0[0] * v_mx + (M0[1] - O1[1]) * v_my + M0[2] * v_mz + Az_c * VS)
    a0g = (M0[0]**2 + (M0[1] - O1[1])**2 + M0[2]**2) - (Ad[0]**2 + (Ad[1] - O1[1])**2 + Az_c**2)
    discg = a1g**2 - 4 * a2g * a0g
    t_geom = td + TL + 999
    if discg >= 0:
        sqrt_discg = np.sqrt(discg)
        for tg in [(-a1g - sqrt_discg) / (2 * a2g), (-a1g + sqrt_discg) / (2 * a2g)]:
            if td < tg < td + TL: t_geom = min(t_geom, tg)

    t_end = td + TL
    t_flight_end = MISSILE_T_FLIGHT.get(missile, 1e9)
    if t_end > t_flight_end:
        t_end = t_flight_end

    # Evaluate initial C2/C1 state
    f0 = f_C2(td)
    c2_on = (f0 is not None and f0 >= 0)
    Mt0 = M_fn(td, missile)
    At0 = Ad - np.array([0., 0., VS * 0])
    c1_on = (np.linalg.norm(Mt0 - At0) <= R_SMOKE)

    # Pair C2 roots
    c2_pairs = []
    c2_idx = 0; cur_start = td; active = c2_on
    while c2_idx < len(c2_roots):
        if active:
            end_t = c2_roots[c2_idx]
            if end_t > cur_start: c2_pairs.append((cur_start, end_t))
            active = False; cur_start = c2_roots[c2_idx]
        else:
            cur_start = c2_roots[c2_idx]; active = True
        c2_idx += 1
    if active:
        end_t = t_end
        if end_t > cur_start: c2_pairs.append((cur_start, end_t))

    # Pair C1 roots
    c1_pairs = []
    c1_idx = 0; cur_start = td; active = c1_on
    while c1_idx < len(c1_roots):
        if active:
            end_t = c1_roots[c1_idx]
            if end_t > cur_start: c1_pairs.append((cur_start, end_t))
            active = False; cur_start = c1_roots[c1_idx]
        else:
            cur_start = c1_roots[c1_idx]; active = True
        c1_idx += 1
    if active:
        end_t = t_end
        if end_t > cur_start: c1_pairs.append((cur_start, end_t))

    # Merge with constraints
    intervals = []
    for (s, e) in c2_pairs:
        valid_end = min(e, t_geom, t_end)
        if valid_end > s: intervals.append((s, valid_end))
    for (s, e) in c1_pairs:
        if e > s: intervals.append((s, e))

    # Fallback: 解析法失败时用步进扫描
    if not intervals:
        dur_start = None
        for tt in np.arange(td, t_end, 0.1):
            Mt = M_fn(tt, missile)
            age = tt - td
            At = Ad - np.array([0., 0., VS * age])
            MA = np.linalg.norm(Mt - At)
            shielded = False
            if MA <= R_SMOKE:
                shielded = True
            else:
                MO1 = np.linalg.norm(Mt - O1)
                AO1 = np.linalg.norm(At - O1)
                if MO1 > AO1:
                    th, al = th_al(tt, Ad, td, missile)
                    if th is not None and th <= al:
                        shielded = True
            if shielded:
                if dur_start is None:
                    dur_start = tt
            else:
                if dur_start is not None:
                    intervals.append((dur_start, tt))
                    dur_start = None
        if dur_start is not None:
            intervals.append((dur_start, t_end))
    return intervals


def union_duration(intervals_list):
    """多个区间列表并集的总长度"""
    all_iv = []
    for ivs in intervals_list:
        all_iv.extend(ivs)
    if not all_iv:
        return 0.0
    all_iv.sort(key=lambda x: x[0])
    merged = [list(all_iv[0])]
    for s, e in all_iv[1:]:
        if s <= merged[-1][1] + 1e-6:
            merged[-1][1] = max(merged[-1][1], e)
        else:
            merged.append([s, e])
    return sum(e - s for s, e in merged)


def multi_bomb_T_eff(bombs, missile='M1'):
    """多枚弹对单导弹的总遮蔽时长(并集)"""
    intervals_list = []
    for Ad, td in bombs:
        ivs = compute_T_eff_intervals(Ad, td, missile)
        intervals_list.append(ivs)
    return union_duration(intervals_list)


def drone_feasible(Ad, td, drone_start=FY1_0, drone_z=1800.):
    """检查无人机能否将炸弹送到Ad并在td起爆"""
    disp = Ad[:2] - drone_start[:2]
    dist_h = np.linalg.norm(disp)
    if td < 1e-6: return None
    v = dist_h / td
    if v < V_MIN or v > V_MAX: return None
    t_fall = np.sqrt(max(0, 2 * (drone_z - Ad[2]) / G))
    t_rel = td - t_fall
    if t_rel < 0: return None
    return {'v': v, 'theta': np.arctan2(disp[1], disp[0]),
            't_release': t_rel, 't_fall': t_fall, 'drone_start': drone_start}


def save_result_xlsx(result, filename):
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"  [warning] openpyxl not installed, can't save {filename}")
        return
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投放策略"
    ws.append(["Parameter", "Value", "Unit"])
    for c in [ws['A1'], ws['B1'], ws['C1']]:
        c.font = Font(bold=True, size=13)

    for k, v in result.get('summary', {}).items():
        ws.append([k, v, ''])

    ws.append([])
    ws.append(["=== Bomb Details ===", "", ""])
    headers = ['BombID', 'Drone', 'Missile', 't_rel(s)', 't_det(s)', 't_delay(s)',
               'Ax(m)', 'Ay(m)', 'Az(m)', 'T_eff(s)', 'v(m/s)', 'theta(deg)']
    ws.append(headers)
    for c in range(1, len(headers)+1):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)

    for b in result.get('bombs', []):
        ws.append([b.get(k, '') for k in [
            'bomb_id', 'drone', 'missile', 't_release', 't_det', 't_delay',
            'A_x', 'A_y', 'A_z', 'T_eff', 'v', 'theta'
        ]])

    ws.append([])
    ws.append(["=== Shielding Analysis ===", "", ""])
    ws.append(['Missile', 'Total T_eff(s)', 'Intervals'])
    for m, info in result.get('shielding', {}).items():
        iv_str = '; '.join([f"[{s:.3f},{e:.3f}]" for s, e in info.get('intervals', [])])
        ws.append([m, round(info.get('duration', 0), 4), iv_str])

    for col in range(1, ws.max_column + 1):
        max_w = max((len(str(ws.cell(r, col).value or '')) for r in range(1, ws.max_row + 1)), default=8)
        ws.column_dimensions[get_column_letter(col)].width = min(max_w + 4, 45)

    wb.save(filename)
    print(f"  [OK] saved to {filename}")
