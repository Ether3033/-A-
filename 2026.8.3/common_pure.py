"""
纯Python版本的核心计算模块 — 使用 math 模块而非 numpy
=========================================================
目的: 绕过 Python 3.13/3.14 + numpy 2.x 在 Windows 上的 C 层 segfault。
所有向量运算用 list/tuple + math 实现，完全无 numpy/scipy 依赖。
"""

import math
import random

# ======================== 场景常量 ========================
M1_0  = (20000.0, 0.0, 2000.0)
M2_0  = (19000.0, 600.0, 2100.0)
M3_0  = (18000.0, -600.0, 1900.0)

FY1_0 = (17800.0, 0.0, 1800.0)
FY2_0 = (12000.0, 1400.0, 1400.0)
FY3_0 = (6000.0, -3000.0, 700.0)
FY4_0 = (11000.0, 2000.0, 1800.0)
FY5_0 = (13000.0, -2000.0, 1300.0)

DRONE_POS = {'FY1': FY1_0, 'FY2': FY2_0, 'FY3': FY3_0,
             'FY4': FY4_0, 'FY5': FY5_0}

O_DECOY = (0.0, 0.0, 0.0)
O1 = (0.0, 200.0, 0.0)

Rc, Hc = 7.0, 10.0
R_SMOKE = 10.0
VS = 3.0
TL = 20.0
VM = 300.0
G = 9.8

V_MIN, V_MAX = 70.0, 140.0

# 导弹飞行总时长 (从发现点到假目标原点)
# T = ||M0 - O_decoy|| / VM
def _dist3(a, b=(0., 0., 0.)):
    return math.sqrt((a[0]-b[0])**2 + (a[1]-b[1])**2 + (a[2]-b[2])**2)

MISSILE_T_FLIGHT = {
    'M1': _dist3(M1_0) / VM,  # ≈ 67.0s
    'M2': _dist3(M2_0) / VM,  # ≈ 63.8s
    'M3': _dist3(M3_0) / VM,  # ≈ 60.4s
}

# 导弹方向单位向量
def _norm3(v):
    d = math.sqrt(v[0]**2 + v[1]**2 + v[2]**2)
    return (v[0]/d, v[1]/d, v[2]/d)

def _sub3(a, b):
    return (a[0]-b[0], a[1]-b[1], a[2]-b[2])

def _add3(a, b):
    return (a[0]+b[0], a[1]+b[1], a[2]+b[2])

def _mul3(v, s):
    return (v[0]*s, v[1]*s, v[2]*s)

def _dot3(a, b):
    return a[0]*b[0] + a[1]*b[1] + a[2]*b[2]

def _len3(v):
    return math.sqrt(v[0]**2 + v[1]**2 + v[2]**2)

def _len2(v):
    return math.sqrt(v[0]**2 + v[1]**2)

u_m1 = _norm3(_sub3(O_DECOY, M1_0))
u_m2 = _norm3(_sub3(O_DECOY, M2_0))
u_m3 = _norm3(_sub3(O_DECOY, M3_0))

MISSILE_PARAMS = {
    'M1': {'M0': M1_0, 'u_m': u_m1},
    'M2': {'M0': M2_0, 'u_m': u_m2},
    'M3': {'M0': M3_0, 'u_m': u_m3},
}


def M_fn(t, missile='M1'):
    p = MISSILE_PARAMS[missile]
    return _add3(p['M0'], _mul3(p['u_m'], VM * t))


def sample_boundary(Mp):
    """纯Python版本: 圆柱体边界采样 (215点)"""
    Mx, My, Mz = Mp
    Ox, Oy, Oz = O1

    d_xy = (Mx - Ox, My - Oy)
    dist_xy = _len2(d_xy)
    if dist_xy < 1e-6:
        return []

    ux, uy = d_xy[0]/dist_xy, d_xy[1]/dist_xy
    psi = math.asin(min(1.0, max(-1.0, Rc / dist_xy)))
    cp, sp = math.cos(psi), math.sin(psi)

    tRx = ux*cp - uy*sp
    tRy = ux*sp + uy*cp
    tLx = ux*cp + uy*sp
    tLy = -ux*sp + uy*cp

    TR = (Ox + Rc*tRx, Oy + Rc*tRy)
    TL_pt = (Ox + Rc*tLx, Oy + Rc*tLy)

    pts = []
    ns, ne = 20, 48

    # 侧面采样点
    for i in range(ns + 1):
        z = Oz + Hc * i / ns
        pts.append((TL_pt[0], TL_pt[1], z))
        pts.append((TR[0], TR[1], z))

    # 顶圆采样点
    for i in range(ne + 1):
        a = 2 * math.pi * i / ne
        pts.append((Ox + Rc*math.cos(a), Oy + Rc*math.sin(a), Oz + Hc))

    # 底面前沿采样点
    aM = math.atan2(My - Oy, Mx - Ox)
    for i in range(ne // 2 + 1):
        a = aM + math.pi/2 + math.pi*i/(ne//2)
        pts.append((Ox + Rc*math.cos(a), Oy + Rc*math.sin(a), Oz))

    # 内点采样
    for i in range(1, 10):
        f = i / 11.0
        Tx = TL_pt[0] + f*(TR[0] - TL_pt[0])
        Ty = TL_pt[1] + f*(TR[1] - TL_pt[1])
        vx = Tx - Ox
        vy = Ty - Oy
        vn = math.sqrt(vx**2 + vy**2)
        if vn < 1e-10:
            continue
        vx *= Rc / vn
        vy *= Rc / vn
        Tx = Ox + vx
        Ty = Oy + vy
        for j in range(ns // 2 + 1):
            z = Oz + Hc * j / (ns // 2)
            pts.append((Tx, Ty, z))

    # 方向向量
    vecs = []
    for p in pts:
        vx = p[0] - Mx
        vy = p[1] - My
        vz = p[2] - Mz
        d = math.sqrt(vx**2 + vy**2 + vz**2)
        if d > 1e-10:
            vecs.append((vx/d, vy/d, vz/d))
    return vecs


def th_al(t, Ad, td, missile='M1'):
    """纯Python版本: 计算theta_max和alpha"""
    Mt = M_fn(t, missile)
    age = t - td
    if age < 0 or age > TL:
        return None, None

    At = (Ad[0], Ad[1], Ad[2] - VS * age)

    dx = Mt[0] - At[0]
    dy = Mt[1] - At[1]
    dz = Mt[2] - At[2]
    MA = math.sqrt(dx**2 + dy**2 + dz**2)

    if MA <= R_SMOKE:
        return 0.0, math.pi / 2

    uAx = (At[0] - Mt[0]) / MA
    uAy = (At[1] - Mt[1]) / MA
    uAz = (At[2] - Mt[2]) / MA

    alpha = math.asin(R_SMOKE / MA)

    proj = sample_boundary(Mt)
    if not proj:
        return None, None

    max_cos = -2.0
    for v in proj:
        cos_val = uAx*v[0] + uAy*v[1] + uAz*v[2]
        if cos_val > max_cos:
            max_cos = cos_val

    max_cos = max(-1.0, min(1.0, max_cos))
    theta_max = math.acos(max_cos)
    return theta_max, alpha


def _bisect_pure(f, a, b, xtol=1e-6, maxiter=50):
    """纯Python二分法"""
    fa, fb = f(a), f(b)
    if fa is None or fb is None:
        raise ValueError("function returned None")
    if not (math.isfinite(fa) and math.isfinite(fb)):
        raise ValueError("function returned non-finite value")
    if fa * fb >= 0:
        raise ValueError("f(a) and f(b) must have opposite signs")
    for _ in range(maxiter):
        c = (a + b) / 2.0
        if abs(b - a) < xtol:
            return c
        fc = f(c)
        if fc is None or not math.isfinite(fc):
            return c
        if fa * fc < 0:
            b, fb = c, fc
        else:
            a, fa = c, fc
    return (a + b) / 2.0


def safe_compute_T_eff(Ad, td, missile='M1'):
    """带异常保护的 compute_T_eff，出错返回 0.0"""
    try:
        return compute_T_eff(Ad, td, missile)
    except Exception:
        return 0.0


def compute_T_eff(Ad, td, missile='M1'):
    """纯Python版本: 计算有效遮蔽时长"""
    p = MISSILE_PARAMS[missile]
    M0 = p['M0']
    u_m = p['u_m']

    def f_C2(t):
        th, al = th_al(t, Ad, td, missile)
        if th is None:
            return -1.0
        return al - th

    # 粗扫描 C2 根
    c2_roots = []
    t_scan = [td + (TL * i / 39) for i in range(40)]

    for i in range(len(t_scan) - 1):
        f1 = f_C2(t_scan[i])
        f2 = f_C2(t_scan[i + 1])
        if f1 is not None and f2 is not None and f1 * f2 < 0:
            try:
                r = _bisect_pure(f_C2, t_scan[i], t_scan[i + 1])
                c2_roots.append(r)
            except Exception:
                pass

    # C1 边界: 二次方程
    dx0 = M0[0] - Ad[0]
    dy0 = M0[1] - Ad[1]
    dz0 = M0[2] - Ad[2] - VS * td
    vx = u_m[0] * VM
    vy = u_m[1] * VM
    vz = u_m[2] * VM + VS
    a2 = vx**2 + vy**2 + vz**2
    a1 = 2 * (dx0 * vx + dy0 * vy + dz0 * vz)
    a0 = dx0**2 + dy0**2 + dz0**2 - R_SMOKE**2
    disc = a1**2 - 4 * a2 * a0
    c1_roots = []
    if disc >= 0:
        sqrt_disc = math.sqrt(disc)
        for tc in [(-a1 - sqrt_disc) / (2 * a2), (-a1 + sqrt_disc) / (2 * a2)]:
            if td < tc < td + TL:
                c1_roots.append(tc)
        c1_roots.sort()

    # 几何失效边界: 二次方程
    v_mx, v_my, v_mz = u_m[0] * VM, u_m[1] * VM, u_m[2] * VM
    Az_c = Ad[2] + VS * td
    a2g = (v_mx**2 + v_my**2 + v_mz**2) - VS**2
    a1g = 2 * (M0[0] * v_mx + (M0[1] - 200) * v_my + M0[2] * v_mz + Az_c * VS)
    a0g = (M0[0]**2 + (M0[1] - 200)**2 + M0[2]**2) - (Ad[0]**2 + (Ad[1] - 200)**2 + Az_c**2)
    discg = a1g**2 - 4 * a2g * a0g
    t_geom = td + TL + 999
    if discg >= 0:
        sqrt_discg = math.sqrt(discg)
        for tg in [(-a1g - sqrt_discg) / (2 * a2g), (-a1g + sqrt_discg) / (2 * a2g)]:
            if td < tg < td + TL:
                t_geom = min(t_geom, tg)

    t_end = td + TL
    # 导弹到达假目标后即失效, 遮蔽时长不超过导弹飞行时间
    t_flight_end = MISSILE_T_FLIGHT.get(missile, 1e9)
    if t_end > t_flight_end:
        t_end = t_flight_end
    intervals = []

    # C2→C1 遮蔽段
    if c2_roots:
        ts = c2_roots[0]
        te = c1_roots[0] if c1_roots else min(t_geom, t_end)
        if te > ts:
            intervals.append((ts, te))

    # C1 球内遮蔽段
    if c1_roots:
        ts = c1_roots[0]
        te = c1_roots[1] if len(c1_roots) > 1 else t_end
        if te > ts:
            intervals.append((ts, te))

    # 退化情况: 步进扫描
    if not intervals:
        dur = 0.0
        steps = int((t_end - td) / 0.1) + 1
        for i in range(steps):
            tt = td + i * 0.1
            if tt > t_end:
                break
            Mt = M_fn(tt, missile)
            age = tt - td
            At = (Ad[0], Ad[1], Ad[2] - VS * age)
            dx = Mt[0] - At[0]
            dy = Mt[1] - At[1]
            dz = Mt[2] - At[2]
            MA = math.sqrt(dx**2 + dy**2 + dz**2)
            if MA <= R_SMOKE:
                dur += 0.1
                continue
            dxo = Mt[0] - O1[0]
            dyo = Mt[1] - O1[1]
            dzo = Mt[2] - O1[2]
            MO1 = math.sqrt(dxo**2 + dyo**2 + dzo**2)
            dxa = At[0] - O1[0]
            dya = At[1] - O1[1]
            dza = At[2] - O1[2]
            AO1 = math.sqrt(dxa**2 + dya**2 + dza**2)
            if MO1 <= AO1:
                continue
            th, al = th_al(tt, Ad, td, missile)
            if th is not None and th <= al:
                dur += 0.1
        return dur

    return sum(e - s for s, e in intervals)


def compute_T_eff_intervals(Ad, td, missile='M1'):
    """纯Python版本: 返回遮蔽区间列表"""
    p = MISSILE_PARAMS[missile]
    M0 = p['M0']
    u_m = p['u_m']

    def f_C2(t):
        th, al = th_al(t, Ad, td, missile)
        if th is None:
            return -1.0
        return al - th

    c2_roots = []
    t_scan = [td + (TL * i / 39) for i in range(40)]
    for i in range(len(t_scan) - 1):
        f1 = f_C2(t_scan[i])
        f2 = f_C2(t_scan[i + 1])
        if f1 is not None and f2 is not None and f1 * f2 < 0:
            try:
                r = _bisect_pure(f_C2, t_scan[i], t_scan[i + 1])
                c2_roots.append(r)
            except Exception:
                pass

    dx0 = M0[0] - Ad[0]
    dy0 = M0[1] - Ad[1]
    dz0 = M0[2] - Ad[2] - VS * td
    vx = u_m[0] * VM
    vy = u_m[1] * VM
    vz = u_m[2] * VM + VS
    a2 = vx**2 + vy**2 + vz**2
    a1 = 2 * (dx0 * vx + dy0 * vy + dz0 * vz)
    a0 = dx0**2 + dy0**2 + dz0**2 - R_SMOKE**2
    disc = a1**2 - 4 * a2 * a0
    c1_roots = []
    if disc >= 0:
        sqrt_disc = math.sqrt(disc)
        for tc in [(-a1 - sqrt_disc) / (2 * a2), (-a1 + sqrt_disc) / (2 * a2)]:
            if td < tc < td + TL:
                c1_roots.append(tc)
        c1_roots.sort()

    v_mx, v_my, v_mz = u_m[0] * VM, u_m[1] * VM, u_m[2] * VM
    Az_c = Ad[2] + VS * td
    a2g = (v_mx**2 + v_my**2 + v_mz**2) - VS**2
    a1g = 2 * (M0[0] * v_mx + (M0[1] - 200) * v_my + M0[2] * v_mz + Az_c * VS)
    a0g = (M0[0]**2 + (M0[1] - 200)**2 + M0[2]**2) - (Ad[0]**2 + (Ad[1] - 200)**2 + Az_c**2)
    discg = a1g**2 - 4 * a2g * a0g
    t_geom = td + TL + 999
    if discg >= 0:
        sqrt_discg = math.sqrt(discg)
        for tg in [(-a1g - sqrt_discg) / (2 * a2g), (-a1g + sqrt_discg) / (2 * a2g)]:
            if td < tg < td + TL:
                t_geom = min(t_geom, tg)

    t_end = td + TL
    # 导弹到达假目标后即失效, 遮蔽时长不超过导弹飞行时间
    t_flight_end = MISSILE_T_FLIGHT.get(missile, 1e9)
    if t_end > t_flight_end:
        t_end = t_flight_end
    intervals = []
    if c2_roots:
        ts = c2_roots[0]
        te = c1_roots[0] if c1_roots else min(t_geom, t_end)
        if te > ts:
            intervals.append((ts, te))
    if c1_roots:
        ts = c1_roots[0]
        te = c1_roots[1] if len(c1_roots) > 1 else t_end
        if te > ts:
            intervals.append((ts, te))
    return intervals


def union_duration(intervals_list):
    """多个区间列表并集的总长度"""
    all_iv = []
    for ivs in intervals_list:
        all_iv.extend(ivs)
    if not all_iv:
        return 0.0
    all_iv.sort(key=lambda x: x[0])
    merged = [list(all_iv[0])]
    for s, e in all_iv[1:]:
        if s <= merged[-1][1] + 1e-6:
            merged[-1][1] = max(merged[-1][1], e)
        else:
            merged.append([s, e])
    return sum(e - s for s, e in merged)


def multi_bomb_T_eff(bombs, missile='M1'):
    """多枚弹对单导弹的总遮蔽时长(并集)"""
    intervals_list = []
    for Ad, td in bombs:
        ivs = compute_T_eff_intervals(Ad, td, missile)
        intervals_list.append(ivs)
    return union_duration(intervals_list)


def drone_feasible(Ad, td, drone_start=FY1_0, drone_z=1800.0):
    """检查无人机可行性"""
    dx = Ad[0] - drone_start[0]
    dy = Ad[1] - drone_start[1]
    dist_h = math.sqrt(dx**2 + dy**2)
    if td < 1e-6:
        return None
    v = dist_h / td
    if v < V_MIN or v > V_MAX:
        return None
    t_fall = math.sqrt(max(0, 2 * (drone_z - Ad[2]) / G))
    t_rel = td - t_fall
    if t_rel < 0:
        return None
    return {'v': v, 'theta': math.atan2(dy, dx),
            't_release': t_rel, 't_fall': t_fall, 'drone_start': drone_start}


def save_result_xlsx(result, filename):
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
    except ImportError:
        print(f"  [warning] openpyxl not installed, can't save {filename}")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投放策略"
    ws.append(["Parameter", "Value", "Unit"])
    for c in [ws['A1'], ws['B1'], ws['C1']]:
        c.font = Font(bold=True, size=13)
    for k, v in result.get('summary', {}).items():
        ws.append([k, v, ''])
    ws.append([])
    ws.append(["=== Bomb Details ===", "", ""])
    headers = ['BombID', 'Drone', 'Missile', 't_rel(s)', 't_det(s)', 't_delay(s)',
               'Ax(m)', 'Ay(m)', 'Az(m)', 'T_eff(s)', 'v(m/s)', 'theta(deg)']
    ws.append(headers)
    for c in range(1, len(headers)+1):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
    for b in result.get('bombs', []):
        ws.append([b.get(k, '') for k in [
            'bomb_id', 'drone', 'missile', 't_release', 't_det', 't_delay',
            'A_x', 'A_y', 'A_z', 'T_eff', 'v', 'theta'
        ]])
    ws.append([])
    ws.append(["=== Shielding Analysis ===", "", ""])
    ws.append(['Missile', 'Total T_eff(s)', 'Intervals'])
    for m, info in result.get('shielding', {}).items():
        iv_str = '; '.join([f"[{s:.3f},{e:.3f}]" for s, e in info.get('intervals', [])])
        ws.append([m, round(info.get('duration', 0), 4), iv_str])
    for col in range(1, ws.max_column + 1):
        max_w = max((len(str(ws.cell(r, col).value or '')) for r in range(1, ws.max_row + 1)), default=8)
        ws.column_dimensions[get_column_letter(col)].width = min(max_w + 4, 45)
    wb.save(filename)
    print(f"  [OK] saved to {filename}")
