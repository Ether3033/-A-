"""
问题3 v2: FY1投放3枚烟幕弹干扰M1 — 系统(v,θ)网格 + BO精化
==========================================================
策略:
  Phase 0: (v,θ)网格搜索 — 对每组(v,θ)枚举炸弹时序配置
  Phase 1: BO精化 — 以网格最优为锚点, GP+EI精细优化8D

核心约束: v,θ 对3弹共享, 投放间隔≥1s, td递增
"""
import sys, os, time, gc, warnings
warnings.filterwarnings('ignore')

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from common import *
import numpy as np
from scipy.linalg import cholesky, solve_triangular
from scipy.stats import norm

T_MAX_SINGLE = compute_T_eff(np.array([17525., 9., 1760.]), 3.007, missile='M1')
G = 9.8

# ======================== 目标函数 ========================
def eval_bomb(v, theta, td, tdl):
    """单弹评估, 返回 (T_single, intervals)"""
    Az = 1800. - 0.5 * G * tdl**2
    if Az < 50 or Az > 1798: return 0.0, []
    tr = td - tdl
    if tr < 0: return 0.0, []
    Ax = FY1_0[0] + v * np.cos(theta) * td
    Ay = FY1_0[1] + v * np.sin(theta) * td
    Ad = np.array([Ax, Ay, Az])
    if drone_feasible(Ad, td) is None: return 0.0, []
    try:
        ivs = compute_T_eff_intervals(Ad, td, 'M1')
        T_s = sum(e - s for s, e in ivs)
        if T_s < 1e-6:
            T_s = compute_T_eff(Ad, td, 'M1')
        return max(0.0, T_s), ivs
    except Exception:
        return 0.0, []

def eval_3bomb(x):
    """8D完整评估: x = [v, theta, td1, tdl1, td2, tdl2, td3, tdl3]"""
    v, theta = float(x[0]), float(x[1])
    if v < V_MIN or v > V_MAX: return -1.0, None
    ivs_list = []; T_list = []; tr_list = []
    for i in range(3):
        td = float(x[2 + 2*i]); tdl = float(x[3 + 2*i])
        if td <= 0.1 or tdl < 0.001: return -1.0, None
        T_s, ivs = eval_bomb(v, theta, td, tdl)
        T_list.append(T_s); ivs_list.append(ivs)
        tr_list.append(td - tdl)
    # 约束
    if not (x[2] < x[4] < x[6]): return -0.5, None
    for i in range(1, 3):
        gap = tr_list[i] - tr_list[i-1]
        if gap < 1.0 - 1e-4: return -(1.0 - gap) - 0.1, None
    # 单弹超界惩罚
    penalty = 0.0
    for T_s in T_list:
        if T_s > T_MAX_SINGLE + 1e-6:
            penalty += 10.0 * (T_s - T_MAX_SINGLE)
    total = union_duration(ivs_list)
    return total - penalty, (total, ivs_list, T_list, tr_list)

def obj_q3(x):
    val, _ = eval_3bomb(x)
    return val

# ======================== Phase 0: (v,θ) 系统网格搜索 ========================
def grid_search_p3(verbose=True):
    """对(v,θ)网格每组枚举炸弹时序, 找全局最优锚点"""
    v_vals = [85, 90, 95, 100, 105]
    th_vals = [176, 178, 180, 182]

    best_union = 0.0; best_x = None; best_detail = None
    n_tested = 0; n_feas = 0

    for v in v_vals:
        for th_deg in th_vals:
            th = np.radians(th_deg)
            # 炸弹时序枚举 (粗粒度)
            for td1 in [2.5, 3.0]:
                for tdl1 in [2.5, 2.85]:
                    tr1 = td1 - tdl1
                    if tr1 < 0: continue
                    T1, ivs1 = eval_bomb(v, th, td1, tdl1)
                    if T1 <= 0: continue

                    for td2 in np.arange(td1 + 0.8, min(td1 + 4.5, 14), 0.8):
                        for tdl2 in [2.5, 3.0, 3.5]:
                            tr2 = td2 - tdl2
                            if tr2 < tr1 + 1.0: continue
                            T2, ivs2 = eval_bomb(v, th, td2, tdl2)
                            if T2 <= 0: continue

                            for td3 in np.arange(td2 + 0.8, min(td2 + 4.5, 16), 0.8):
                                for tdl3 in [2.8, 3.3, 3.8]:
                                    tr3 = td3 - tdl3
                                    if tr3 < tr2 + 1.0: continue
                                    T3, ivs3 = eval_bomb(v, th, td3, tdl3)
                                    if T3 <= 0: continue

                                    n_tested += 1
                                    total = union_duration([ivs1, ivs2, ivs3])
                                    if total > 0: n_feas += 1
                                    if total > best_union:
                                        best_union = total
                                        best_x = np.array([v, th, td1, tdl1, td2, tdl2, td3, tdl3])
                                        best_detail = (ivs1, ivs2, ivs3, T1, T2, T3)
    if verbose:
        print(f"  (v,θ)网格: {len(v_vals)}×{len(th_vals)}={len(v_vals)*len(th_vals)}组")
        print(f"  测试 {n_tested} 个配置, {n_feas} 个有效")
        if best_x is not None:
            ivs1, ivs2, ivs3, T1, T2, T3 = best_detail
            print(f"  最优: v={best_x[0]:.0f} θ={np.degrees(best_x[1]):.1f}° Union={best_union:.4f}s")
            print(f"    Bomb1: td={best_x[2]:.3f} tdl={best_x[3]:.3f} T={T1:.3f}s")
            print(f"    Bomb2: td={best_x[4]:.3f} tdl={best_x[5]:.3f} T={T2:.3f}s")
            print(f"    Bomb3: td={best_x[6]:.3f} tdl={best_x[7]:.3f} T={T3:.3f}s")
    return best_x, best_union, best_detail

# ======================== BO 组件 (8D) ========================
def rbf_kernel_batch(X1, X2, ls, sigma2):
    X1_s = X1 / ls[None, :]; X2_s = X2 / ls[None, :]
    sqdist = (np.sum(X1_s**2, axis=1)[:, None]
              + np.sum(X2_s**2, axis=1)[None, :]
              - 2 * X1_s @ X2_s.T)
    return sigma2 * np.exp(-0.5 * np.clip(sqdist, 0, None))

class GP:
    def __init__(self, ls, sigma2, noise=1e-6):
        self.ls = np.asarray(ls); self.sigma2 = sigma2; self.noise = noise
        self.X = None; self.y = None; self.L = None; self.alpha = None
        self._ym = 0.0; self._ys = 1.0

    def fit(self, X, y):
        self.X = np.asarray(X); raw = np.asarray(y)
        self._ym = np.mean(raw); self._ys = max(np.std(raw), 1e-8)
        self.y = (raw - self._ym) / self._ys
        n = len(self.X)
        K = rbf_kernel_batch(self.X, self.X, self.ls, self.sigma2)
        K += np.eye(n) * self.noise
        for attempt in range(3):
            try:
                self.L = cholesky(K, lower=True)
                break
            except np.linalg.LinAlgError:
                K += np.eye(n) * (10 ** (-6 + attempt*2))
        self.alpha = solve_triangular(self.L.T,
                        solve_triangular(self.L, self.y, lower=True))

    def predict_batch(self, X_pred, batch_size=2000):
        X_pred = np.asarray(X_pred); n_pred = len(X_pred)
        mu = np.zeros(n_pred); sigma = np.zeros(n_pred)
        for start in range(0, n_pred, batch_size):
            end = min(start + batch_size, n_pred); Xb = X_pred[start:end]
            K_s = rbf_kernel_batch(Xb, self.X, self.ls, self.sigma2)
            v = solve_triangular(self.L, K_s.T, lower=True)
            mu[start:end] = K_s @ self.alpha
            sigma[start:end] = np.sqrt(np.maximum(
                self.sigma2 - np.sum(v**2, axis=0), 1e-12))
        return mu * self._ys + self._ym, sigma * self._ys

def maximize_ei(gp, y_best, lb, ub, n_candidates=15000):
    d = len(lb)
    X_cand = np.random.uniform(lb, ub, (n_candidates, d))
    mu, sigma = gp.predict_batch(X_cand, batch_size=3000)
    sigma = np.maximum(sigma, 1e-12)
    delta = mu - y_best - 0.01
    Z = delta / sigma
    ei = delta * norm.cdf(Z) + sigma * norm.pdf(Z)
    return X_cand[np.argmax(ei)], np.max(ei)

def bayesian_optimize(obj_fn, lb, ub, n_init=20, n_iter=80, seed=42,
                       known_points=None, verbose=True):
    np.random.seed(seed)
    d = len(lb)
    X = np.zeros((n_init, d))
    for j in range(d):
        perm = np.random.permutation(n_init)
        X[:, j] = lb[j] + (ub[j] - lb[j]) * (perm + np.random.uniform(size=n_init)) / n_init
    if known_points:
        for i, (xp, _) in enumerate(known_points):
            if i < n_init: X[i] = xp
    # 强制td递增
    for i in range(n_init):
        td = sorted([X[i, 2], X[i, 4], X[i, 6]])
        X[i, 2], X[i, 4], X[i, 6] = td[0], td[1], td[2]

    y = np.array([obj_fn(x) for x in X])
    best_idx = np.argmax(y); y_best = y[best_idx]; x_best = X[best_idx].copy()
    n_feas_init = np.sum(y > 0)
    if verbose:
        print(f"  LHS {n_init}pts | feas={n_feas_init}/{n_init} | best={y_best:.4f}s")

    gp = GP(ls=(ub - lb) * 0.2, sigma2=2.0, noise=1e-3)
    t0 = time.time()
    for it in range(1, n_iter + 1):
        gp.fit(X, y)
        x_next, ei_val = maximize_ei(gp, y_best, lb, ub)
        y_next = obj_fn(x_next)
        X = np.vstack([X, x_next]); y = np.append(y, y_next)
        if y_next > y_best: y_best = y_next; x_best = x_next.copy()
        if verbose and (it % 25 == 0 or it == 1 or it == n_iter):
            n_feas = np.sum(y > 0)
            print(f"  iter {it:3d}/{n_iter} | EI={ei_val:.4f} | "
                  f"best={y_best:.4f}s | feas={n_feas}/{len(y)} | {time.time()-t0:.1f}s")
        if it % 30 == 0: gc.collect()

    if verbose: print(f"  BO done: {len(X)} evals | best={y_best:.4f}s")
    return x_best, y_best

# ======================== 辅助函数 ========================
def extract_bombs_info(x_opt, v_opt, th_opt):
    bombs_info = []
    for i in range(3):
        td = float(x_opt[2 + 2*i]); tdl = float(x_opt[3 + 2*i])
        Ax = FY1_0[0] + v_opt * np.cos(th_opt) * td
        Ay = FY1_0[1] + v_opt * np.sin(th_opt) * td
        Az = 1800. - 0.5 * G * tdl**2
        Ad = np.array([Ax, Ay, Az])
        tr = td - tdl
        T_s = compute_T_eff(Ad, td, 'M1')
        ivs = compute_T_eff_intervals(Ad, td, 'M1')
        bombs_info.append({
            'Ad': Ad, 't_det': td, 't_delay': tdl,
            't_release': tr, 'T_eff': T_s, 'intervals': ivs
        })
    return bombs_info

def save_and_print(bombs_info, v_opt, th_opt, label, x_opt=None):
    """打印结果并保存到xlsx"""
    all_intervals = [bi['intervals'] for bi in bombs_info]
    total_u = union_duration(all_intervals)

    print(f"\n{'='*70}")
    print(f"{label}")
    print(f"{'='*70}")
    print(f"  Total T_eff (union): {total_u:.4f}s")
    print(f"  Drone: v={v_opt:.2f} m/s  theta={np.degrees(th_opt):.4f} deg")
    print(f"  {'Bomb':>5s} {'t_rel':>8s} {'t_det':>8s} {'t_delay':>8s} "
          f"{'Ax':>9s} {'Ay':>8s} {'Az':>7s} {'T_single':>9s}  Intervals")
    print(f"  {'-'*85}")
    for i, bi in enumerate(bombs_info):
        iv_str = '; '.join([f"[{s:.3f},{e:.3f}]" for s, e in bi['intervals']])
        print(f"  {i+1:5d} {bi['t_release']:8.4f} {bi['t_det']:8.4f} "
              f"{bi['t_delay']:8.4f} {bi['Ad'][0]:9.1f} {bi['Ad'][1]:8.1f} "
              f"{bi['Ad'][2]:7.1f} {bi['T_eff']:9.4f}  {iv_str}")
    print(f"\n  P2 optimal: {T_MAX_SINGLE:.4f}s | P3 union: {total_u:.4f}s ({total_u/T_MAX_SINGLE:.2f}x P2)")
    if x_opt is not None:
        print(f"  Config: v={x_opt[0]:.2f} θ={np.degrees(x_opt[1]):.4f}° "
              f"td=[{x_opt[2]:.3f},{x_opt[4]:.3f},{x_opt[6]:.3f}] "
              f"tdl=[{x_opt[3]:.3f},{x_opt[5]:.3f},{x_opt[7]:.3f}]")

    # 保存xlsx
    try:
        proj_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xlsx_path = os.path.join(proj_root, '附件', 'result1.xlsx')
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb['Sheet1']
        for i, bi in enumerate(bombs_info):
            row = 2 + i
            ws.cell(row=row, column=2,  value=round(np.degrees(th_opt), 4))
            ws.cell(row=row, column=3,  value=round(v_opt, 2))
            rx = FY1_0[0] + v_opt * np.cos(th_opt) * bi['t_release']
            ry = FY1_0[1] + v_opt * np.sin(th_opt) * bi['t_release']
            ws.cell(row=row, column=5,  value=round(rx, 1))
            ws.cell(row=row, column=6,  value=round(ry, 1))
            ws.cell(row=row, column=7,  value=round(1800.0, 1))
            ws.cell(row=row, column=8,  value=round(bi['Ad'][0], 1))
            ws.cell(row=row, column=9,  value=round(bi['Ad'][1], 1))
            ws.cell(row=row, column=10, value=round(bi['Ad'][2], 1))
            ws.cell(row=row, column=11, value=round(bi['T_eff'], 4))
            ws.cell(row=row, column=12, value='M1')
        wb.save(xlsx_path)
        print(f"  [OK] Saved to {xlsx_path}")
    except Exception as e:
        print(f"  [WARN] xlsx failed: {e}")
    return total_u

# ======================== 主程序 ========================
if __name__ == "__main__":
    t_start = time.time()
    print("=" * 70)
    print("Problem 3 v2: FY1 x 3 bombs -> M1 | (v,θ) Grid + BO")
    print(f"P2 single-bomb optimum: {T_MAX_SINGLE:.4f}s")
    print("=" * 70)

    # ---- Phase 0: 多锚点构造 (P2最优 + 已知BO结果 + 网格最优) ----
    print(f"\n{'─'*70}")
    print("Phase 0: 多锚点构造")
    print(f"{'─'*70}")

    # 锚点1: P2最优 v,θ (Bomb1最优, Bomb2/3=0)
    anchor1_x = np.array([91.5, np.radians(178.13), 3.007, 2.857, 6.0, 3.5, 9.0, 4.0])
    anchor1_y, _ = eval_3bomb(anchor1_x)
    print(f"  Anchor1 (P2 v,θ): Union={anchor1_y:.4f}s")

    # 锚点2: 之前BO找到的最优 (v=94.61, θ=178.68°)
    anchor2_x = np.array([94.61, np.radians(178.68), 2.979, 2.882, 5.351, 3.577, 7.275, 3.029])
    anchor2_y, _ = eval_3bomb(anchor2_x)
    print(f"  Anchor2 (prev BO): Union={anchor2_y:.4f}s")

    # 选最好的锚点
    anchors = [(anchor1_x, anchor1_y), (anchor2_x, anchor2_y)]
    anchors.sort(key=lambda a: -a[1])
    best_anchor_x, best_anchor_y = anchors[0]

    v_a, th_a = float(best_anchor_x[0]), float(best_anchor_x[1])
    bombs_anchor = extract_bombs_info(best_anchor_x, v_a, th_a)
    save_and_print(bombs_anchor, v_a, th_a, "PHASE 0: BEST ANCHOR", x_opt=best_anchor_x)

    # ---- Phase 1: BO精化 (多轮, 逐步缩小搜索框) ----
    print(f"\n{'─'*70}")
    print("Phase 1: BO (GP+EI) 精细优化 — 3轮逐步收敛")
    print(f"{'─'*70}")

    xc = best_anchor_x.copy()
    current_best = best_anchor_y

    for round_idx, (scale, n_init, n_iter) in enumerate([
        (2.5, 20, 60),   # 第1轮: 宽搜索
        (1.2, 15, 40),   # 第2轮: 中搜索
        (0.5, 10, 25),   # 第3轮: 窄精化
    ]):
        print(f"\n  --- Round {round_idx+1}: scale={scale}, n_init={n_init}, n_iter={n_iter} ---")

        delta = np.array([10.0, np.radians(8.0), 3.0, 2.0, 5.0, 2.5, 6.0, 3.0])
        lb_bo = np.maximum(xc - delta * scale, [
            70.0, np.radians(160.0), 0.5, 0.05, 2.0, 0.5, 3.0, 0.5])
        ub_bo = np.minimum(xc + delta * scale, [
            140.0, np.radians(200.0), 8.0, 5.5, 14.0, 6.0, 16.0, 6.5])

        print(f"    v[{lb_bo[0]:.0f},{ub_bo[0]:.0f}] θ[{np.degrees(lb_bo[1]):.0f},{np.degrees(ub_bo[1]):.0f}]°")
        print(f"    td1[{lb_bo[2]:.1f},{ub_bo[2]:.1f}] td2[{lb_bo[4]:.1f},{ub_bo[4]:.1f}] td3[{lb_bo[6]:.1f},{ub_bo[6]:.1f}]")

        gc.collect(); time.sleep(0.2)
        try:
            x_bo, y_bo = bayesian_optimize(
                obj_q3, lb_bo, ub_bo, n_init=n_init, n_iter=n_iter,
                seed=42 + round_idx*10,
                known_points=[(xc, current_best)], verbose=True)
            gc.collect()

            if y_bo > current_best + 0.001:
                xc = x_bo.copy()
                current_best = y_bo
                print(f"    >> Improved: {current_best:.4f}s")
            else:
                print(f"    >> No significant improvement")
        except Exception as e:
            print(f"    [WARN] Round {round_idx+1} crashed: {e}")

    # ---- 最终结果 ----
    v_final, th_final = float(xc[0]), float(xc[1])
    bombs_final = extract_bombs_info(xc, v_final, th_final)
    final_union = save_and_print(bombs_final, v_final, th_final, "FINAL RESULT", x_opt=xc)

    elapsed = time.time() - t_start
    print(f"\n{'='*70}")
    print(f"Problem 3 v2 complete! ({elapsed:.1f}s total)")
    print(f"{'='*70}")
